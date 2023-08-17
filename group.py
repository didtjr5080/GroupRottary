from openpyxl import load_workbook
import random
import  datetime as dt
import os
import re

f = open("excelname.txt","r",encoding='utf-8')
line=f.readline()
print("엑셀 파일 이름:",line)
load_xlsx = load_workbook(line,data_only=True) #엑셀 파일 이름
f.close()
load_xlsx = load_workbook("room.xlsx",data_only=True) #엑셀 파일 이름

f = open("loadsheet.txt","r",encoding='utf-8')
line=f.readline()
print("시트 이름:",line)
load_sheet = load_xlsx[line] #시트 이름
f.close()

all_people=[] #엑셀에서 불러온 모든 인원

success_count=0 #성공적으로 추첨한 인원

pick_list = [] #뽑은 리스트


execpt_list = [] #제외할 명단


f = open("maxcount.txt",'r')
line = f.readline()
max_count = int(line) #조당 최대 인원
print("조당 최대 인원:" ,max_count)
f.close()




# f = open("all_people.txt",'r')
# line = f.readline()
# all_people_num = int(line) #반 전체 인원 (제외할 인원수 뺀 인원수)
# print("반 전체 인원:" ,all_people_num)
# f.close()



group_1 = [] #1조 리스트
group_2 = [] #2조 리스트
group_3 = [] #3조 리스트
group_4 = [] #4조 리스트
group_5 = [] #5조 리스트
group_6 = [] #6조 리스트
group_7 = [] #7조 리스트
group_8 = [] #8조 리스트

f = open("startsheet.txt",'r')
line = f.readline()
start_sheet = line #시작 시트
print("시작 시트:" ,start_sheet)
f.close()

f = open("endsheet.txt",'r')
line = f.readline()
end_sheet = line #마지막 시트
print("마지막 시트:" ,end_sheet)
f.close()

ragne_cell = load_sheet[start_sheet:end_sheet]
for row in ragne_cell:
    for cell in row:
        # print(cell.value)
        all_people.append(cell.value)

f = open("exeptlist.txt",'rt',encoding='utf-8')
lines = f.readlines()
for line in lines:
    # print(line)
    line = line.replace("\n","")
    line = line.replace(" ", "")
    # print(all_people[all_people.index(lines)])
    if line in all_people:
        execpt_list.append(line)
    else:
        print("유효하지 않은 제외명단입니다."," 이름:",line)



# print(execpt_list)
execpt_list_print=""
for i in execpt_list:
    execpt_list_print = execpt_list_print + "\n" + i
    # print(execpt_list_print)
print("제외할 명단:",execpt_list_print)
f.close()


f = open("all_people.txt",'wt', encoding= 'utf-8')
all_people_num = len(all_people)-len(execpt_list) #총인워수에서 제외할 인원수 뺸 숫자
f.write(str(all_people_num))
print(all_people_num)

# def suc_pick(group_num,pick):
#     print(group_num, "조:", pick)
#     pick_list.append(pick)

# print(all_people)

while success_count != all_people_num:
    pick = random.choice(all_people)
    # print(pick)
    # print(pick_list)
    if pick in execpt_list:
        continue
    else:
        if pick not in pick_list:
            group_num = random.randint(1, 8)
            # print(group_num)
            if group_num == 1:
                if len(group_1) >= max_count:
                    continue
                else:
                    group_1.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1

            elif group_num == 2:
                if len(group_2) >= max_count:
                    continue
                else:
                    group_2.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1
            elif group_num == 3:
                if len(group_3) >= max_count:
                    continue
                else:
                    group_3.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1
            elif group_num == 4:
                if len(group_4) >= max_count:
                    continue
                else:
                    group_4.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1
            elif group_num == 5:
                if len(group_5) >= max_count:
                    continue
                else:
                    group_5.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1
            elif group_num == 6:
                if len(group_6) >= max_count:
                    continue
                else:
                    group_6.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1
            elif group_num == 7:
                if len(group_7) >= max_count:
                    continue
                else:
                    group_7.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1
            elif group_num == 8:
                if len(group_8) >= max_count:
                    continue
                else:
                    group_8.append(pick)
                    # print(group_num, "조:", pick)
                    pick_list.append(pick)
                    success_count += 1
        elif pick in pick_list:
            continue

#리스트로 저장돼 있는 조 추첨 결과 가공
g1_text="1조:"
for i in group_1:
    g1_text = g1_text + i +" "
    # print(g1_text)

g2_text="2조:"
for i in group_2:
    g2_text = g2_text + i +" "
    # print(g2_text)

g3_text="3조:"
for i in group_3:
    g3_text = g3_text + i +" "
    # print(g3_text)

g4_text="4조:"
for i in group_4:
    g4_text = g4_text + i +" "
    # print(g4_text)

g5_text="5조:"
for i in group_5:
    g5_text = g5_text + i +" "
    # print(g5_text)

g6_text="6조:"
for i in group_6:
    g6_text = g6_text + i +" "
    # print(g6_text)

g7_text="7조:"
for i in group_7:
    g7_text = g7_text + i +" "
    # print(g7_text)

g8_text="8조:"
for i in group_8:
    g8_text = g8_text + i +" "
    # prignt(g8_text)

#조 추첨 결과 프린트
print(g1_text)
print(g2_text)
print(g3_text)
print(g4_text)
print(g5_text)
print(g6_text)
print(g7_text)
print(g8_text)

# now_time = str((dt.d  atetime.now().year,dt.datetime.now().month,dt.datetime.now().day,dt.datetime.now().hour,dt.datetime.now().minute,dt.datetime.now().second))
now_time = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
print("추첨 결과 파일 이름: "+now_time+".txt")
# now_time_set =""
# for i in now_time:
#     now_time_set = str(now_time)+i
#     print(now_time_set)
# print(now_time)

current_path = os.getcwd() #현재 디랙토리 불러오기
# print(current_path + r"\result") 

try:
    f = open(current_path + r"\result\\"  + now_time + ".txt", 'x') # 현재 디렉토리/result 디렉토리 열기 시도
except:
    os.mkdir("result") #만약 현재 디렉토리/result 디렉토리 열기 시도를 실패 한다면 result 폴더 생성
    # 조 추첨 결과 저장
    f = open(current_path + r"\result\\" + now_time + ".txt", 'x')  # 현재 디렉토리/result 디렉토리 열기 시도
    for i in range(1,5):
        f.write(g1_text)
    # f.write(g1_text)
    # f.write('\n')
    # f.write(g2_text)
    # f.write('\n')
    # f.write(g3_text)
    # f.write('\n')
    # f.write(g4_text)
    # f.write('\n')
    # f.write(g5_text)
    # f.write('\n')
    # f.write(g6_text)
    # f.write('\n')
    # f.write(g7_text)
    # f.write('\n')
    # f.write(g8_text)
    # f.write('\n')
    # f.close()
else:
    #조 추첨 결과 저장
    f = open(current_path + r"\result\\" + now_time + ".txt", 'x')  # 현재 디렉토리/result 디렉토리 열기 시도

    for i in range(1,5):
        f.write(g1_text)
    # f.write('\n')
    # f.write(g2_text)
    # f.write('\n')
    # f.write(g3_text)
    # f.write('\n')
    # f.write(g4_text)
    # f.write('\n')
    # f.write(g5_text)
    # f.write('\n')
    # f.write(g6_text)
    # f.write('\n')
    # f.write(g7_text)
    # f.write('\n')
    # f.write(g8_text)
    # f.write('\n')
    # f.close()

os.system('pause')
# print("1조",group_1)
# print("2조",group_2)
# print("3조",group_3)
# print("4조",group_4)
# print("5조",group_5)
# print("6조",group_6)
# print("7조",group_7)
# print("8조",group_8)


